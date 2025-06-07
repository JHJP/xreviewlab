from flask import Flask, render_template, request, session, send_file, redirect, url_for, jsonify
import pandas as pd
from optimizer_gurobi import optimize, METHODS, compute_damage
import numpy as np
# 추가: 워드클라우드 관련
from collections import Counter
import io, base64
from wordcloud import WordCloud
from konlpy.tag import Okt
import matplotlib.pyplot as plt
import os
from dotenv import load_dotenv
import openai

# ───────────────────────────────────────────────────────────────
# 📊  Brand-reputation research → ₩ per DAMAGE-UNIT
# ----------------------------------------------------------------
# Why 12 500 ₩ ?  (sources ⬇️)
# • Harvard Business School (Luca 2016): every ★ on Yelp shifts revenue
#   by ≈ 5-9 %. 
# • Gominga “Online Review Stats 2024”: one bad review can scare away
#   ~30 shoppers & 40 % of prospects. 
# • Statistics Korea (2024): avg e-commerce basket ≈ ₩51 800.
#
# Worst-case loss for a *high-damage* review:
#     30 customers × ₩51 800  ≈ ₩1 554 000
# The same review’s `compute_damage()` score peaks ≈ 125 units
# (explicit warning + long text + aged).                ⟶  1 unit
#     1 554 000 / 125  ≈ ₩12 400  → rounded → **₩12 500**
# ───────────────────────────────────────────────────────────────
WON_PER_DAMAGE_UNIT: float = 12_500

def risk_bucket(score: float, low_thresh: float, high_thresh: float):
    """
    Return (slug, korean_label)  e.g.  ( "medium", "보통" )
    * low_thresh  : at-or-below this   → low   (낮음)
    * high_thresh : at-or-below this   → medium(보통)
    * above high_thresh               → high  (높음)
    """
    if score <= low_thresh:
        return ("low",    "낮음")
    elif score <= high_thresh:
        return ("medium", "보통")
    else:
        return ("high",   "높음")

app = Flask(__name__)
app.secret_key = 'replace_this_with_a_random_secret_1234'

# 진행상황 초기화
def init_progress(total):
    session['progress_current'] = 0
    session['progress_total'] = total
    session.modified = True

def increment_progress():
    session['progress_current'] = session.get('progress_current', 0) + 1
    session.modified = True

def get_progress():
    return session.get('progress_current', 0), session.get('progress_total', 1)

@app.route('/progress')
def progress():
    current, total = get_progress()
    print('[progress API]', current, total)  # 서버 콘솔에 직접 찍기
    return jsonify({'current': current, 'total': total})

@app.route("/", methods=["GET", "POST"])
def index():
    # ── 1. show welcome/tutorial overlay only the first time ───────────────
    if request.method == "POST":
        session["tutorial_seen"] = True 
    tutorial_mode = not session.get("tutorial_seen", False)

    # ── 2. skeletons for the template (stay None on plain GET)
    recs = metrics = insights = None
    wordcloud_url = None

    # ── 3. Handle the POST: run the optimiser & build the dashboard
    if request.method == "POST":

        # ── 3-a. grab inputs ─────────────────────────────────────────────────
        budget          = float(request.form["budget"])
        hours_available = float(request.form["total_hours_available"])

        # 브랜드 이름 받아서 코드 추출
        brand_name = request.form.get("brand_name")
        brand_code = None
        if brand_name:
            try:
                from crawlers.olive_crawler import get_brand_code, crawl_brand_all, crawl_reviews_for_goods
                brand_code = get_brand_code(brand_name)
                total_products_df = crawl_brand_all(brand_code, limit=3)
                goods_nos = total_products_df['goodsNo'].tolist()
                total_reviews_df = crawl_reviews_for_goods(goods_nos, limit=5)
            except Exception as e:
                print(f"[ERROR] 올리브영 브랜드코드 추출 실패: {e}")
            prodlist_df = crawl_brand_all(brand_code)
        else:
            df              = pd.read_csv("reviews_sample_15.csv")

        # optional “allowed methods”; default = all five
        allowed = request.form.getlist("methods") or METHODS

        # ── 3-b. optimise & build dashboard numbers ──────────────────────────
        rec_df, sens = optimize(
            df,
            budget,
            hours_available,
            allowed_methods=allowed
        )
        init_progress(len(df))  # 전체 리뷰 개수로 초기화 (POST에서만!)

        # ── 3-c. add risk buckets for UI chips ─────────────────────────────
        # percentiles ≈ terciles → dynamic to the current batch
        p33, p66 = np.percentile(rec_df.risk_score, [33, 66])
        rec_df[["risk_slug", "risk_label"]] = rec_df.risk_score.apply(
            lambda s: pd.Series(risk_bucket(s, p33, p66))
        )
        # ──────────────────────────────────────────────────────────────
        # 3-c.  What-if insight cards (from LP sensitivity)
        # ──────────────────────────────────────────────────────────────
        # Shadow prices:  ( ≤ constraints ⇒ Pi ≤ 0 in a minimisation model )
        shadow = sens.get("shadow_prices") or {}
        rhs    = shadow.get("rhs_ranges", {})
        sp     = shadow.get("shadow_prices", {})
        budget_slack = sens.get("budget_slack", None)
        time_slack = sens.get("time_slack", None)
        save_per_1k_budget = float(sp.get("budget", 0)) * 1000
        save_per_hour = float(sp.get("time", 0))
        budget_note = None
        time_note = None
        if save_per_1k_budget == 0 and budget_slack is not None and budget_slack > 0:
            budget_note = "예산 제약이 느슨(slack)해져 한계효과(Shadow price)가 0입니다. 예산을 줄이거나 α 불확실성을 낮추면 값이 나올 수 있습니다."
        if save_per_hour == 0 and time_slack is not None and time_slack > 0:
            time_note = "시간 제약이 느슨(slack)해져 한계효과(Shadow price)가 0입니다. 시간을 줄이거나 α 불확실성을 낮추면 값이 나올 수 있습니다."
        insights = {
            "save_per_1k_budget": save_per_1k_budget,
            "save_per_hour":     save_per_hour,
            "budget_headroom":   float(rhs.get("budget", (None, None))[1] or 0) - float(budget),
            "hour_headroom":     float(rhs.get("time", (None, None))[1] or 0) - float(hours_available),
            "budget_note": budget_note,
            "time_note": time_note,
        }
        # Convert to customer-friendly “additional damage PREVENTED”
        whatif = dict(
            save_per_1k_budget = max(0.0, -float(sp.get("budget", 0)) * WON_PER_DAMAGE_UNIT * 1_000),   # ₩ saved / ₩1 000
            save_per_hour      = max(0.0, -float(sp.get("time", 0)) * WON_PER_DAMAGE_UNIT),             # ₩ saved / hour
            # Optional head-room: how much budget could still yield that return
            budget_headroom = max(0.0,
                float(rhs.get("budget", (None, None))[1] or 0) - float(budget)
                if "rhs_ranges" in sens and "budget" in sens["rhs_ranges"] and isinstance(sens["rhs_ranges"]["budget"], (list, tuple)) and len(sens["rhs_ranges"]["budget"]) > 1 else 0.0
            ),
            hour_headroom = max(0.0,
                sens["rhs_ranges"]["time"][1] - hours_available
                if "rhs_ranges" in sens and "time" in sens["rhs_ranges"] and isinstance(sens["rhs_ranges"]["time"], (list, tuple)) and len(sens["rhs_ranges"]["time"]) > 1 else 0.0
            ),
        )
        # ────────────────  🎯  HARD-ROI METRICS  ────────────────
        # 1.   예 상 손 해 방 지 금 액  =  {(총 damage) – (MILP 목적함수)}*WON_PER_DAMAGE_UNIT
        damage_before = compute_damage(df).sum()
        damage_after  = sens.get("obj_worst_case")  # solver objective
        if damage_after is None:
            # Fallback: show error message to user and skip calculation
            return render_template(
                "index.html",
                tutorial_mode=tutorial_mode,
                methods=METHODS,
                recommended_responses=None,
                metrics=None,
                insights=None,
                error_msg="최적화 결과에서 목적함수 값을 찾을 수 없습니다. 입력값을 확인하거나 관리자에게 문의하세요."
            )
        net_damage_units = damage_before - damage_after          # ▲units
        net_damage_prevented = net_damage_units * WON_PER_DAMAGE_UNIT

        # 2.   예 상 별 점 상 승 량
        if "review_stars" in df.columns:
            old_avg = df.review_stars.mean()

            # optimistic assumption: 대응되면 최소 ★4로 회복
            new_stars = df.review_stars.copy()
            treated_ids = set(rec_df.review_id)
            low_mask = (df.index.isin(treated_ids)) & (df.review_stars < 4)
            new_stars.loc[low_mask] = 4

            avg_star_delta = float(new_stars.mean() - old_avg)
        else:
            avg_star_delta = 0.0

        # 3.   예 상 응 답 시 간 절 감 량
        TIME_PER_METHOD = {        # hours per review after automation
            "text_reply":      0.0833,   # 5 min (5/60)
            "chat_followup":   0.1667,   # 10 min (10/60)
            "phone_call":      0.5000,   # 30 min
            "discount_coupon": 0.0833,
            "vip_apology":     0.2500,
        }
        baseline_manual_time = 0.75          # 45 min(45min per review)
        baseline_total_hours = len(df) * baseline_manual_time

        optimised_total_hours = sum(
            TIME_PER_METHOD.get(row.method, 0.0) for row in rec_df.itertuples()
        )
        response_time_saved = baseline_total_hours - optimised_total_hours

        # 4.   브 랜 드 평 판 회 복 률  (= 낮은 별점 중 몇 %를 구해냈나)
        if "review_stars" in df.columns:
            negative_mask = df.review_stars < 4
            total_neg = int(negative_mask.sum())
            recovered = int(
                df.loc[negative_mask & df.index.isin(treated_ids)].shape[0]
            )
            review_recovery_rate = recovered / total_neg if total_neg else 0.0
        else:
            review_recovery_rate = 0.0

        metrics = dict(
            net_damage_prevented = net_damage_prevented,
            avg_star_delta       = avg_star_delta,
            response_time_saved  = response_time_saved,
            review_recovery_rate = review_recovery_rate,
        )


        # 각 리뷰별 키워드 추출 (한글/영어 자동 지원)
        import re
        def extract_keywords(text):
            text = str(text)
            # 한글 포함 여부
            if re.search(r"[\uac00-\ud7a3]", text):
                # 한글 리뷰: konlpy 사용
                okt = Okt()
                return [w for w in okt.nouns(text) if len(w) > 1]
            else:
                # 영어 리뷰: nltk 사용, punkt_tab 오류 대응
                import nltk
                try:
                    nltk.data.find('tokenizers/punkt')
                except LookupError:
                    nltk.download('punkt')
                try:
                    nltk.data.find('corpora/stopwords')
                except LookupError:
                    nltk.download('stopwords')
                try:
                    from nltk.corpus import stopwords
                    from nltk.tokenize import word_tokenize
                    words = word_tokenize(text)
                except LookupError:
                    try:
                        nltk.download('punkt_tab')
                        from nltk.tokenize import word_tokenize
                        words = word_tokenize(text)
                    except Exception:
                        # 마지막 fallback: 공백 기준 분리
                        words = text.split()
                stop_words = set(nltk.corpus.stopwords.words('english'))
                return [w.lower() for w in words if w.isalpha() and w.lower() not in stop_words and len(w) > 2]

        def extract_insight_keywords(review_text):
            import os, json
            from openai import OpenAI
            api_key = os.getenv('OPENAI_API_KEY')
            if not api_key:
                return extract_keywords(review_text)

            client = OpenAI(api_key=api_key)

            few_shot_messages = [
                {"role": "user",
                "content": "리뷰: 직원이 친절했지만 음식이 너무 늦게 나왔어요."},
                {"role": "assistant",
                "content": '{"keywords":["직원 친절","음식 지연"]}'},
                {"role": "user",
                "content": "리뷰: 가격이 비싸고 품질이 기대 이하여서 실망했습니다."},
                {"role": "assistant",
                "content": '{"keywords":["가격","품질"]}'},
            ]

            # 1) 키워드 추출 ---------------------------------------------------------
            system_keyword = (
                "너는 비즈니스 의사결정을 돕는 분석가다. "
                "아래 리뷰에서 핵심 키워드 3~6개를 추출한다. "
                "각 키워드는 한국어 명사구여야 한다. "
                "출력은 다음 JSON 스키마를 반드시 따른다.\n\n"
                '{\n  "keywords": ["키워드1", "키워드2", ...]\n}\n'
                "다른 텍스트(접두어·주석·마침표 등)를 절대 포함하지 마라."
            )
            user_keyword = f"리뷰: {review_text}"

            messages= (
                    [{"role": "system", "content": system_keyword}]
                    + few_shot_messages
                    + [{"role": "user", "content": user_keyword}]
                )
            
            try:
                resp = client.chat.completions.create(
                    model="gpt-4o",
                    messages=messages,
                    response_format={"type": "json_object"},
                    temperature=0.0,
                    max_tokens=128,
                )

                data = json.loads(resp.choices[0].message.content)
                keywords = data.get("keywords", [])
                if not isinstance(keywords, list):
                    raise ValueError("`keywords` 필드 형식 오류")
                return keywords

            except Exception:  # 파싱 실패 · API 오류 등
                return extract_keywords(review_text)




        def gpt_classify_and_reply(review_text):
            import os, json
            from openai import OpenAI
            api_key = os.getenv('OPENAI_API_KEY')
            if not api_key:
                return ("분류불가", "AI 답변 생성에 실패했습니다. (API KEY 없음/오류)")

            client = OpenAI(api_key=api_key)

            # 1) 분류 -----------------------------------------------------------------
            system_cat = (
                "너는 고객 리뷰를 단일 불만 유형으로 분류하는 AI 모델이다. "
                "반드시 아래 JSON 스키마를 지켜서 한국어로 출력해라.\n\n"
                "스키마:\n"
                "{\n  \"category\": \"품질 | 가격 | 서비스 | 기타\"\n}\n"
                "※ 가능한 값은 딱 네 가지 중 하나다. 그 외 단어·마침표·공백을 넣지 마라."
            )
            user_cat = f"리뷰: {review_text}"

            cat_resp = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": system_cat},
                    {"role": "user", "content": user_cat}
                ],
                response_format={"type": "json_object"},  # JSON mode
                temperature=0.0,
                max_tokens=16,
            )

            try:
                category = json.loads(cat_resp.choices[0].message.content)["category"]
            except Exception:
                category = "분류실패"

            # 2) 답변 -----------------------------------------------------------------
            system_reply = (
                "너는 고객응대 직원(agent)이다. 다음 지침을 지켜 2~4문장 한국어로 답하라.\n"
                "① 진심 어린 사과를 1문장 포함한다.\n"
                "② 실제 가능한 내부 조치(원인 점검, 재발 방지 교육 등)만 제시한다.\n"
                "③ 금전·쿠폰·사은품·무료 혜택 약속은 절대 언급하지 않는다.\n"
                "④ 서비스 회복 패러독스(사후 대처가 만족을 높일 수 있다는 개념)를 염두에 두되, 현학적 용어는 쓰지 마라.\n"
                "⑤ 회사·브랜드명은 넣지 않는다."
            )
            user_reply = f"리뷰: {review_text}"

            reply_resp = client.chat.completions.create(
                model="gpt-4o",
                messages=[
                    {"role": "system", "content": system_reply},
                    {"role": "user", "content": user_reply}
                ],
                temperature=0.6,
                max_tokens=256,
            )

            suggested_reply = reply_resp.choices[0].message.content.strip()
            return (category, suggested_reply)

        recs = []
        for idx, row in enumerate(df.to_dict('records')):
            rec_row = next((r for r in rec_df.to_dict('records') if int(r['review_id']) == int(row.get('review_id', idx))), None)
            if rec_row is not None:
                kws = extract_insight_keywords(rec_row['review_text'])
                category, suggested_reply = gpt_classify_and_reply(rec_row['review_text'])
                # 날짜는 반드시 원본 df 기준
                review_date = row.get('date') or row.get('날짜') or row.get('DATE') or ''
                # 답변 실패 안내 메시지 처리
                if suggested_reply.strip() in ["(답변실패)", "", None]:
                    print(f"[GPT 답변실패] idx={idx}, review_id={rec_row.get('review_id')}, text={rec_row['review_text']}")
                    suggested_reply = "AI 답변 생성에 실패했습니다. 관리자에게 문의하세요."
                rec = {
                    'id': idx,
                    'review_id': rec_row['review_id'],
                    'method': rec_row['method'],
                    'risk_score_norm': rec_row['risk_score_norm'],
                    'review_text': rec_row['review_text'],
                    'keywords': kws,
                    'category': category,
                    'suggested_reply': suggested_reply,
                    'date': review_date
                }
                recs.append(rec)
                increment_progress()  # GPT 응답 시에만 카운트 증가
        # 반드시 리다이렉트로 세션 보존 (추천조치 결과도 세션에 저장)
        session['recs'] = recs
        session.modified = True
        # index 함수 정상 종료: render_template 등으로 반환

    return render_template(
        "index.html",
        tutorial_mode          = tutorial_mode,
        methods                = METHODS,
        recommended_responses  = recs,
        metrics                = metrics,
        insights               = insights,
        wordcloud_url          = wordcloud_url
    )





import tempfile
from io import BytesIO
from PIL import Image
import base64
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage

@app.route("/download_report", methods=["GET"])
def download_report():
    report_data = session.get('report_data')
    if not report_data:
        return redirect(url_for('index'))
    recs = report_data['recs']
    metrics = report_data['metrics']
    insights = report_data['insights']
    wordcloud_url = report_data['wordcloud_url']
    # 1. 추천조치 테이블 DataFrame 생성
    df = pd.DataFrame(recs)
    # 2. 엑셀 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "추천조치"
    # 헤더
    ws.append(list(df.columns))
    # 데이터
    for row in df.itertuples(index=False):
        ws.append(list(row))
    # 3. 워드클라우드 이미지 삽입 (있을 때만)
    if wordcloud_url and wordcloud_url.startswith('data:image'):
        img_data = base64.b64decode(wordcloud_url.split(',')[1])
        with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp_img:
            tmp_img.write(img_data)
            tmp_img.flush()
            img = XLImage(tmp_img.name)
            img.width = 400
            img.height = 180
            ws.add_image(img, 'H2')
    # 4. metrics, insights 시트 추가
    if metrics:
        ws2 = wb.create_sheet("효과요약")
        for k, v in metrics.items():
            ws2.append([k, v])
    if insights:
        ws3 = wb.create_sheet("인사이트")
        for k, v in insights.items():
            ws3.append([k, v])
    # 5. 파일로 저장 후 다운로드
    with BytesIO() as output:
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="review_report.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

import smtplib
from email.message import EmailMessage

@app.route("/send_report_email", methods=["POST"])
def send_report_email():
    report_data = session.get('report_data')
    if not report_data:
        return redirect(url_for('index'))
    recs = report_data['recs']
    metrics = report_data['metrics']
    insights = report_data['insights']
    wordcloud_url = report_data['wordcloud_url']
    # 1. 엑셀 리포트 생성 (메모리)
    df = pd.DataFrame(recs)
    wb = Workbook()
    ws = wb.active
    ws.title = "추천조치"
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))
    if wordcloud_url and wordcloud_url.startswith('data:image'):
        img_data = base64.b64decode(wordcloud_url.split(',')[1])
        with tempfile.NamedTemporaryFile(delete=False, suffix='.png') as tmp_img:
            tmp_img.write(img_data)
            tmp_img.flush()
            img = XLImage(tmp_img.name)
            img.width = 400
            img.height = 180
            ws.add_image(img, 'H2')
    if metrics:
        ws2 = wb.create_sheet("효과요약")
        for k, v in metrics.items():
            ws2.append([k, v])
    if insights:
        ws3 = wb.create_sheet("인사이트")
        for k, v in insights.items():
            ws3.append([k, v])
    with BytesIO() as output:
        wb.save(output)
        output.seek(0)
        xlsx_bytes = output.read()
    # 2. 이메일 발송
    email_to = request.form.get('email')
    email_user = os.getenv('EMAIL_USER')
    email_password = os.getenv('EMAIL_PASSWORD')
    if not (email_to and email_user and email_password):
        return "이메일 발송 정보가 부족합니다. 관리자에게 문의하세요.", 400
    msg = EmailMessage()
    msg['Subject'] = "리뷰 대응 리포트"
    msg['From'] = email_user
    msg['To'] = email_to
    msg.set_content("첨부파일로 리뷰 대응 리포트를 보내드립니다.")
    msg.add_attachment(xlsx_bytes, maintype='application', subtype='vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename='review_report.xlsx')
    try:
        with smtplib.SMTP_SSL('smtp.gmail.com', 465) as smtp:
            smtp.login(email_user, email_password)
            smtp.send_message(msg)
        return "이메일이 성공적으로 발송되었습니다!", 200
    except Exception as e:
        return f"이메일 발송 실패: {e}", 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)
