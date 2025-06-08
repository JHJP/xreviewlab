from flask import Blueprint, session, redirect, url_for, send_file
import pandas as pd
from openpyxl import Workbook
from io import BytesIO

report_bp = Blueprint('report', __name__)

@report_bp.route("/download_report", methods=["GET"])
def download_report():
    report_data = session.get('report_data')
    if not report_data:
        return redirect(url_for('index'))
    recs = report_data['recs']
    metrics = report_data['metrics']
    insights = report_data['insights']
    wordcloud_url = report_data['wordcloud_url']
    # 1. 추천조치 테이블 DataFrame 생성
    df = pd.DataFrame(recs)
    # 2. 엑셀 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "추천조치"
    ws.append(list(df.columns))
    for row in df.itertuples(index=False):
        ws.append(list(row))
    # 4. metrics, insights 시트 추가
    if metrics:
        ws2 = wb.create_sheet("효과요약")
        for k, v in metrics.items():
            ws2.append([k, v])
    if insights:
        ws3 = wb.create_sheet("인사이트")
        for k, v in insights.items():
            ws3.append([k, v])
    # 5. 파일로 저장 후 다운로드
    with BytesIO() as output:
        wb.save(output)
        output.seek(0)
        return send_file(output, as_attachment=True, download_name="review_report.xlsx", mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
